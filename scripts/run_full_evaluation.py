"""
Full evaluation script — combines:
  1. Semantic map position extraction (perceived vs ground truth)
  2. Experiment runner: SIMPLE (5) + NEGATIVE (5) = 10 commands
  3. Saves everything into a single combined JSON + appends to Excel

Categories:
  simple   — objects that EXIST, basic name commands
  negative — objects that DON'T EXIST (wrong color/shape combos + totally absent)
"""

import json
import re
import csv
import math
import os
import sys
import openpyxl
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))
from semantic_navigation.semantic_map import SemanticMap
from semantic_navigation.mission_controller import MissionController
import numpy as np

# ── Config ───────────────────────────────────────────────────────────────────
_SCRIPTS_DIR = Path(__file__).parent
MAP_PATH   = str(_SCRIPTS_DIR.parent / "data" / "semantic_maps" / "latest.json")
OUTPUT_DIR = str(_SCRIPTS_DIR / "data" / "experiments")
EXCEL_PATH = None  # set to an xlsx path to enable Excel export, e.g. "/path/to/results.xlsx"

GROUND_TRUTH = {
    "red cylinder":   (-2.00,  2.00),
    "blue box":       ( 2.00,  2.00),
    "yellow box":     ( 2.50, -2.50),
    "white cylinder": ( 2.50,  0.50),
    "green cylinder": (-2.00, -2.00),
}
LABEL_REMAP = {
    "blue cube":  "blue box",
    "red pillar": "red cylinder",
    "pillar":     "white cylinder",
}

# ── Test commands ─────────────────────────────────────────────────────────────

# SIMPLE: objects that EXIST — system should find and navigate to them
SIMPLE_COMMANDS = [
    ("Go to the red cylinder",         True, "red cylinder"),
    ("Find the blue box",              True, "blue box"),
    ("Navigate to the green cylinder", True, "green cylinder"),
    ("Go to the yellow box",           True, "yellow box"),
    ("Find the white cylinder",        True, "white cylinder"),
]

# NEGATIVE: objects that DO NOT EXIST — system should reject all of these
NEGATIVE_COMMANDS = [
    # Confusion cases — color exists but wrong shape
    ("Go to the green box",             False, None),
    ("Find the red box",                False, None),
    ("Navigate to the blue cylinder",   False, None),
    ("Go to the yellow cylinder",       False, None),
    # Absent cases — color/object doesn't exist at all
    ("Find the purple object",          False, None),
]


def load_map_robust(path):
    with open(path) as f:
        content = f.read()
    content = content.replace("{{", "{").replace("}}", "}")
    content = re.sub(r",\s*,", ",", content)
    data, _ = json.JSONDecoder().raw_decode(content)
    return data


def extract_perceived_positions(map_data):
    """Get best (highest obs) perceived position per object."""
    best = {}
    for o in map_data["objects"]:
        label = LABEL_REMAP.get(o["label"], o["label"])
        if label not in GROUND_TRUTH:
            continue
        if label not in best or o["observations"] > best[label]["observations"]:
            best[label] = dict(o)
            best[label]["canonical_label"] = label
    return best


def compute_localization_stats(best):
    stats = {}
    for label, gt_pos in GROUND_TRUTH.items():
        if label in best:
            o = best[label]
            px, py = o["position"][0], o["position"][1]
            gx, gy = gt_pos
            err = math.sqrt((px - gx) ** 2 + (py - gy) ** 2)
            stats[label] = {
                "perceived_x":  round(px, 3),
                "perceived_y":  round(py, 3),
                "gt_x":         gx,
                "gt_y":         gy,
                "error_m":      round(err, 3),
                "observations": o["observations"],
                "confidence":   round(o.get("confidence", 0), 3),
                "detected":     True,
            }
        else:
            stats[label] = {
                "perceived_x":  None,
                "perceived_y":  None,
                "gt_x":         GROUND_TRUTH[label][0],
                "gt_y":         GROUND_TRUTH[label][1],
                "error_m":      None,
                "observations": 0,
                "confidence":   0,
                "detected":     False,
            }
    return stats


def run_experiments(smap):
    controller = MissionController(semantic_map=smap, detector=None)

    all_commands = (
        [("simple",   cmd, exp, gt) for cmd, exp, gt in SIMPLE_COMMANDS]
        + [("negative", cmd, exp, gt) for cmd, exp, gt in NEGATIVE_COMMANDS]
    )

    results = []
    current_cat = None

    print(f"\nRunning {len(all_commands)} experiments...\n")

    for category, command, expected_found, gt_label in all_commands:
        if category != current_cat:
            current_cat = category
            print(f"\n--- {category.upper()} ---")

        goal, parsed = controller.execute_command(command)
        actually_found = goal is not None

        if expected_found:
            correct = actually_found
            outcome = "FOUND" if correct else "MISSED"
        else:
            correct = not actually_found
            if correct:
                outcome = "CORRECTLY REJECTED"
            else:
                if goal is not None:
                    best_label, best_dist = None, float("inf")
                    for lbl, gt_pos in GROUND_TRUTH.items():
                        d = np.linalg.norm(np.array(goal[:2]) - np.array(gt_pos))
                        if d < best_dist:
                            best_dist, best_label = d, lbl
                    outcome = (f"CONFUSED → went to {best_label}"
                               if best_dist < 1.5 else "FALSE POSITIVE")
                else:
                    outcome = "FALSE POSITIVE"

        loc_err = None
        if actually_found and gt_label and gt_label in GROUND_TRUTH:
            gx, gy = GROUND_TRUTH[gt_label]
            loc_err = round(float(np.linalg.norm(
                np.array(goal[:2]) - np.array([gx, gy]))), 3)

        result = {
            "command":            command,
            "category":           category,
            "expected_found":     expected_found,
            "actually_found":     actually_found,
            "correct":            correct,
            "outcome":            outcome,
            "target_object":      parsed.query_text,
            "goal_position":      goal.tolist() if goal is not None else None,
            "localization_error": loc_err,
        }
        results.append(result)

        err_str = f" err={loc_err:.2f}m" if loc_err is not None else ""
        tick = "✓" if correct else "✗"
        print(f"  {tick} {command:55} → {outcome}{err_str}")

        controller.state = controller.state.__class__("idle")

    return results


def compute_experiment_metrics(results):
    total   = len(results)
    correct = sum(1 for r in results if r["correct"])
    fp      = sum(1 for r in results if not r["expected_found"] and r["actually_found"])
    fn      = sum(1 for r in results if r["expected_found"] and not r["actually_found"])
    confused = sum(1 for r in results if "CONFUSED" in r["outcome"])

    by_cat = {}
    for cat in ["simple", "negative"]:
        cr = [r for r in results if r["category"] == cat]
        if not cr:
            continue
        cc = sum(1 for r in cr if r["correct"])
        errs = [r["localization_error"] for r in cr
                if r["localization_error"] is not None]

        if cat == "negative":
            confusion_cmds = [c[0] for c in NEGATIVE_COMMANDS[:4]]  # first 4
            absent_cmds    = [c[0] for c in NEGATIVE_COMMANDS[4:]]  # last 1
            confusion_results = [r for r in cr if r["command"] in confusion_cmds]
            absent_results    = [r for r in cr if r["command"] in absent_cmds]
            by_cat[cat] = {
                "total":    len(cr),
                "correct":  cc,
                "accuracy": round(cc / len(cr), 3),
                "avg_localization_error": round(float(np.mean(errs)), 3) if errs else None,
                "confusion_subcategory": {
                    "total":   len(confusion_results),
                    "correct": sum(1 for r in confusion_results if r["correct"]),
                },
                "absent_subcategory": {
                    "total":   len(absent_results),
                    "correct": sum(1 for r in absent_results if r["correct"]),
                },
            }
        else:
            by_cat[cat] = {
                "total":    len(cr),
                "correct":  cc,
                "accuracy": round(cc / len(cr), 3),
                "avg_localization_error": round(float(np.mean(errs)), 3) if errs else None,
            }

    tp_errs = [r["localization_error"] for r in results
               if r["localization_error"] is not None]
    return {
        "total_commands":         total,
        "total_correct":          correct,
        "overall_accuracy":       round(correct / total, 3) if total else 0,
        "false_positives":        fp,
        "false_negatives":        fn,
        "confusion_errors":       confused,
        "avg_localization_error": round(float(np.mean(tp_errs)), 3) if tp_errs else None,
        "by_category":            by_cat,
    }


def append_to_excel(loc_stats, exp_metrics, excel_path):
    def g(label):
        s = loc_stats[label]
        if s["detected"]:
            return (s["perceived_x"], s["perceived_y"], s["confidence"], s["error_m"])
        return ("N/A", "N/A", "N/A", "N/A")

    rc = g("red cylinder")
    yb = g("yellow box")
    bb = g("blue box")
    gc = g("green cylinder")
    wc = g("white cylinder")

    detected_errs = [s["error_m"] for s in loc_stats.values()
                     if s["error_m"] is not None]
    avg_loc_err = round(sum(detected_errs) / len(detected_errs), 2) if detected_errs else "N/A"
    overall_acc = f"{exp_metrics['overall_accuracy']:.1%}"

    # Row: [run#, timestamp, rc_x, rc_y, rc_conf, rc_err, yb_x, yb_y, yb_conf, yb_err, ...]
    row = ["", ""] + list(rc) + list(yb) + list(bb) + list(gc) + list(wc) \
          + [avg_loc_err, overall_acc]

    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        next_row = ws.max_row + 1
        for r in ws.iter_rows(min_row=3):
            if all(c.value is None for c in r[:2]):
                next_row = r[0].row
                break
        for i, val in enumerate(row):
            ws.cell(row=next_row, column=i+1, value=val)
        wb.save(excel_path)
        print(f"\nAppended to Excel row {next_row} — "
              f"avg_loc_err={avg_loc_err}m  overall={overall_acc}")
    except Exception as e:
        print(f"\nCould not write to Excel: {e}")


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # ── 1. Load map ──────────────────────────────────────────────────────────
    print("Loading semantic map...")
    map_data  = load_map_robust(MAP_PATH)
    best      = extract_perceived_positions(map_data)
    loc_stats = compute_localization_stats(best)

    print(f"\nTotal raw objects in map: {len(map_data['objects'])}")
    print(f"\n{'Object':20} {'Perceived (x,y)':22} {'Conf':8} {'GT (x,y)':22} {'Error':8} Obs")
    print("-" * 90)
    for label in sorted(GROUND_TRUTH.keys()):
        s = loc_stats[label]
        if s["detected"]:
            print(f"{label:20} ({s['perceived_x']:+.2f}, {s['perceived_y']:+.2f})"
                  f"           {s['confidence']:.3f}"
                  f"   ({s['gt_x']:+.2f}, {s['gt_y']:+.2f})"
                  f"           {s['error_m']:.2f}m   {s['observations']}")
        else:
            print(f"{label:20} NOT DETECTED")

    # ── 2. Run experiments ───────────────────────────────────────────────────
    smap = SemanticMap()
    smap.load(MAP_PATH)
    exp_results = run_experiments(smap)
    exp_metrics = compute_experiment_metrics(exp_results)

    # Print summary
    print(f"\n{'='*60}")
    print("EXPERIMENT RESULTS")
    print(f"{'='*60}")
    print(f"Overall: {exp_metrics['total_correct']}/{exp_metrics['total_commands']} "
          f"({exp_metrics['overall_accuracy']:.1%})")
    print(f"False positives:  {exp_metrics['false_positives']}")
    print(f"False negatives:  {exp_metrics['false_negatives']}")
    print(f"Confusion errors: {exp_metrics['confusion_errors']}")
    if exp_metrics["avg_localization_error"]:
        print(f"Avg localization error: {exp_metrics['avg_localization_error']:.3f}m")
    print(f"\nBy category:")
    for cat, d in exp_metrics["by_category"].items():
        err = (f", avg_err={d['avg_localization_error']:.2f}m"
               if d.get("avg_localization_error") else "")
        print(f"  {cat:>10}: {d['correct']}/{d['total']} ({d['accuracy']:.1%}){err}")
        if cat == "negative":
            conf = d["confusion_subcategory"]
            abst = d["absent_subcategory"]
            print(f"              ├ confusion (wrong shape): "
                  f"{conf['correct']}/{conf['total']}")
            print(f"              └ absent (doesn't exist):  "
                  f"{abst['correct']}/{abst['total']}")

    # ── 3. Save combined JSON + CSV ──────────────────────────────────────────
    combined = {
        "timestamp":         timestamp,
        "map_path":          MAP_PATH,
        "total_map_objects": len(map_data["objects"]),
        "localization": {
            "per_object": loc_stats,
            "avg_error_m": round(
                sum(s["error_m"] for s in loc_stats.values()
                    if s["error_m"] is not None)
                / max(1, sum(1 for s in loc_stats.values()
                             if s["error_m"] is not None)), 3),
        },
        "experiments": {
            "results": exp_results,
            "metrics": exp_metrics,
        },
    }

    json_path = os.path.join(OUTPUT_DIR, f"full_eval_{timestamp}.json")
    csv_path  = os.path.join(OUTPUT_DIR, f"full_eval_{timestamp}.csv")

    with open(json_path, "w") as f:
        json.dump(combined, f, indent=2)

    with open(csv_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=exp_results[0].keys())
        writer.writeheader()
        writer.writerows(exp_results)

    print(f"\nSaved combined results to:\n  {json_path}\n  {csv_path}")

    # ── 4. Append to Excel (optional) ────────────────────────────────────────
    if EXCEL_PATH:
        append_to_excel(loc_stats, exp_metrics, EXCEL_PATH)


if __name__ == "__main__":
    main()
